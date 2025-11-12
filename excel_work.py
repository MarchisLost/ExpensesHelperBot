import os
import io
from dotenv import load_dotenv
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from openpyxl import load_workbook, utils

# Scopes required for accessing Google Drive
SCOPES = ['https://www.googleapis.com/auth/drive']


def authenticate_google_drive():
    """
    Authenticate and return Google Drive service object.
    Uses credentials.json for authentication.

    Returns:
        googleapiclient.discovery.Resource: Authenticated Drive service
    """
    creds = None

    # Check if we have a token.json file (created after first auth)
    if os.path.exists('token.json'):
        try:
            creds = Credentials.from_authorized_user_file('token.json', SCOPES)
        except Exception as e:
            print(f"Error loading token.json: {e}")
            print("Will re-authenticate...")
            creds = None

    # If no valid credentials, authenticate
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                print("Token expired, refreshing...")
                creds.refresh(Request())
                print("Token refreshed successfully!")
            except Exception as e:
                print(f"Failed to refresh token: {e}")
                print("Re-authenticating...")
                creds = None

        if not creds:
            # Check if credentials.json exists
            if not os.path.exists('credentials.json'):
                raise FileNotFoundError(
                    "credentials.json not found. Please:\n"
                    "1. Download credentials.json from Google Cloud Console\n"
                    "2. Place it in the same directory as this script"
                )

            # Authenticate using credentials.json
            print("Opening browser for authentication...")
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
            print("Authentication successful!")

            # Save credentials for next run
            with open('token.json', 'w') as token:
                token.write(creds.to_json())
            print("Credentials saved to token.json")

    return build('drive', 'v3', credentials=creds)


def download_file_by_id(file_id, destination_path, service=None):
    """
    Download a file from Google Drive by its file ID.

    Args:
        file_id (str): Google Drive file ID
        destination_path (str): Local path where file will be saved
        service: Optional pre-authenticated Drive service object

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        if service is None:
            service = authenticate_google_drive()

        # Get file metadata
        file_metadata = service.files().get(fileId=file_id).execute()
        file_name = file_metadata.get('name', 'unknown_file')

        print(f"Downloading: {file_name}")

        # Request file content
        request = service.files().get_media(fileId=file_id)

        # Create file stream
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)

        done = False
        while done is False:
            status, done = downloader.next_chunk()
            print(f"Download progress: {int(status.progress() * 100)}%")

        # Write to file
        with open(destination_path, 'wb') as f:
            f.write(fh.getvalue())

        print(f"File downloaded successfully to: {destination_path}")
        return True

    except Exception as e:
        print(f"Error downloading file: {str(e)}")
        return False


def get_file_for_editing(file_id, local_filename=None, service=None):
    """
    Download a file from Google Drive for local editing.

    Args:
        file_id (str): Google Drive file ID
        local_filename (str): Optional local filename (defaults to original name)
        service: Optional pre-authenticated Drive service object

    Returns:
        str: Path to the downloaded file, or None if failed
    """
    try:
        if service is None:
            service = authenticate_google_drive()

        # Get file metadata first
        file_metadata = service.files().get(fileId=file_id).execute()
        original_name = file_metadata.get('name', f'file_{file_id}')

        # Use provided filename or original name
        if local_filename is None:
            local_filename = original_name

        print(f"Getting file: {original_name}")

        # Download the file
        success = download_file_by_id(file_id, local_filename, service)

        if success:
            print(f"File ready for editing: {local_filename}")
            return local_filename
        else:
            return None

    except Exception as e:
        print(f"Error getting file: {str(e)}")
        return None


def save_file_back_to_drive(file_id, local_filename, service=None):
    """
    Save your edited local file back to Google Drive (updates the same file ID).

    Args:
        file_id (str): Google Drive file ID to update
        local_filename (str): Path to your edited local file
        service: Optional pre-authenticated Drive service object

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        if service is None:
            service = authenticate_google_drive()

        success = update_file_content(file_id, local_filename, service)

        if success:
            print(f"File saved back to Google Drive! File ID: {file_id}")

        return success

    except Exception as e:
        print(f"Error saving file back to Drive: {str(e)}")
        return False


def update_file_content(file_id, new_file_path, service=None):
    """
    Update the content of an existing file in Google Drive.

    Args:
        file_id (str): Google Drive file ID to update
        new_file_path (str): Local path to the new content
        service: Optional pre-authenticated Drive service object

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        if service is None:
            service = authenticate_google_drive()

        media = MediaFileUpload(new_file_path)
        updated_file = service.files().update(
            fileId=file_id,
            media_body=media
        ).execute()

        print(f"File content updated successfully. File ID: {updated_file.get('id')}")
        return True

    except Exception as e:
        print(f"Error updating file: {str(e)}")
        return False


def edit_file_workflow(file_id, service=None):
    """
    Complete workflow: download file, let you edit it, then save it back.

    Args:
        file_id (str): Google Drive file ID
        service: Optional pre-authenticated Drive service object

    Returns:
        str: Path to local file for editing
    """
    print("=" * 50)
    print("FILE EDITING WORKFLOW")
    print("=" * 50)

    # Download the file
    local_file = get_file_for_editing(file_id, service=service)

    if local_file:
        print(f"\n✅ File downloaded: {local_file}")
        # print(f"save_file_back_to_drive('{file_id}', '{local_file}')")
        print("\n" + "=" * 50)

        return local_file
    else:
        print("❌ Failed to download file")
        return None


def read_cells(file_path: str, sheet_name: str, start_col: int, last_column: int):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]

    # From C to N line 16, 24, 50, 39
    row_number = [16, 24, 39, 52]
    total = 0

    # Create for loop to go through the column letters and add all the values
    for col in range(start_col, last_column):
        for row in row_number:
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value:
                print(f"Row/Col: {row}-{col} and cell value: {cell_value}")
                total += cell_value

    #TODO Call the write function to write the final value
    print(type(total))
    wb.close()
    return total


# def write_cell(file_path:str, sheet_name: str, cell: str, value):
#     #TODO Get the value to write and write
#     wb = load_workbook("file_path")
#     sheet = wb[""]

#     sheet[cell] = value  # write value into the cell

#     wb.save(FILE_PATH)
#     wb.close()


def get_last_month(file_path: str, sheet_name: str):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]

    # Loop through columns C(3) to N(14)
    row_number = 16
    for col in range(3, 15):  # Start from C
        cell_value = sheet.cell(row=row_number, column=col).value

        if cell_value is None or str(cell_value).strip() == "0":
            break  # Stop at first empty column

        # This right now is useless because i want the number of the column and not the letter
        # last_column = utils.get_column_letter(col)
        # print(f"{last_column}{row_number}: {cell_value}")

    return col


def main():
    load_dotenv()
    file_id = os.getenv("FILE_ID")
    sheet_1 = os.getenv("SHEET_1")
    sheet_2 = os.getenv("SHEET_2")

    # Get starting month/column, to transform month into corresponding number
    month_to_col = {
        'jan': 3,
        'fev': 4,
        'feb': 4,
        'mar': 5,
        'abr': 6,
        'apr': 6,
        'mai': 7,
        'may': 7,
        'jun': 8,
        'jul': 9,
        'ago': 10,
        'aug': 10,
        'set': 11,
        'sep': 11,
        'out': 12,
        'oct': 12,
        'nov': 13,
        'dez': 14,
        'dec': 14
    }

    # Get file
    local_file = edit_file_workflow(file_id)
    print(local_file)

    # Check last month with data to get the end month of calculations
    last_month = get_last_month(local_file, sheet_1)
    # print(last_month)

    # Get input of month to start from user, call functions to get the values, calcualte them and print result
    #TODO change the input from this function to discord command
    try:
        user_month = input("Enter starting month (Jan, Fev, Mar...): ").lower().strip()
        start_col = month_to_col.get(user_month)
    except Exception as e:
        print(f"Invalid month with error: {e}")
    else:
        print(f'Valid month {user_month}!')

        # Get the values from the first person
        value_1 = read_cells(local_file, sheet_1, start_col, last_month)
        print(value_1)

        # Get the values from the second person
        value_2 = read_cells(local_file, sheet_2, start_col, last_month)
        print(value_2)

        # Calculate who spent more and how much owes the other person
        final_value = round(value_1 - value_2, 2)
        print(f'Final value: {final_value}')

        if final_value > 0:
            print(f'Person 2 owes Person 1: {final_value}e')
        elif final_value < 0:
            print(f'Person 1 owes Person 2: {final_value}e')
    #TODO return a sentence saying who owes who as well
    #TODO change the way to get the month to get it from discord command
    return final_value


if __name__ == "__main__":
    main()
